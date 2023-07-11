import os.path

from django.shortcuts import render
from django.views import View
import xlrd
import openpyxl
from openpyxl import load_workbook


class TranslateView(View):
    def get(request):

        wb = load_workbook(filename='original.xlsx')

        sheet_ranges = wb['CPU']

        sheet = [
            [sheet_ranges.cell(row=row, column=column).value for column in range(1, 10)]
            for row in range(1, 150)]

        context = {'sheet': sheet}

        return render(request, 'translate/index.html', context=context)
